from csv import reader as csv_reader
from re import sub
from typing import List

import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


def CustomQuit(msg: str) -> None:
    """Выход из программы с выводом сообщения на консоль.
        :param msg: Сообщение, выводимое на консоль.
        """
    print(msg)
    quit()


class CSV:
    """Класс для чтения и обработки CSV-файлов.
        Attributes
        ----------
        data : csv.reader
            Данные, полученные после прочтения CSV-файла при помощи функции reader библиотеки csv.
        title : list
            Список заголовков столбцов CSV-файла.
        rows : list
            Список строк с данными о вакансии. 1 строка = 1 вакансия.
        """
    data: csv_reader
    title: list
    rows: list

    def __init__(self, file_name: str):
        """Инициализирует объект CSV, пытается прочесть файл с переданным именем. Обрабатывает случаи пустого файла и
                отсутствия данных в файле.
                :param file_name: Путь до CSV-файла.
                """
        with open(file_name, 'r', newline='', encoding='utf-8-sig') as file:
            self.data = csv_reader(file)
            try:
                self.title = next(self.data)
            except StopIteration:
                CustomQuit('Пустой файл')

            self.rows = [row for row in self.data
                         if len(list(filter(lambda word: word != '', row))) == len(self.title)]

            if len(self.rows) == 0:
                CustomQuit('Нет данных')


class Translator:
    """Класс для перевода валюты из международного формата на русский язык и с русского языка в числовой формат по
            заранее установленному курсу.
        """
    AZN: str = "Манаты"
    BYR: str = "Белорусские рубли"
    EUR: str = "Евро"
    GEL: str = "Грузинский лари"
    KGS: str = "Киргизский сом"
    KZT: str = "Тенге"
    RUR: str = "Рубли"
    UAH: str = "Гривны"
    USD: str = "Доллары"
    UZS: str = "Узбекский сум"
    currency_to_rub: {str, float} = {
        "Манаты": 35.68,
        "Белорусские рубли": 23.91,
        "Евро": 59.90,
        "Грузинский лари": 21.74,
        "Киргизский сом": 0.76,
        "Тенге": 0.13,
        "Рубли": 1,
        "Гривны": 1.64,
        "Доллары": 60.66,
        "Узбекский сум": 0.0055,
    }

    def Translate(self, key: str, dict_name: str = None) -> str:
        """Переводит из международного формата на русский язык. Если было передано имя словаря, возвращает из него
                значение по ключу.
                :param key: Международный формат валюты или название словаря.
                :param dict_name: Имя словаря, существующего в аттрибутах класса (на данный момент нет доступных).
                """
        if dict_name is not None:
            return self.__getattribute__(dict_name)[key]
        return self.__getattribute__(key)

    def TranslateCurrencyToRub(self, currency: str) -> int or float:
        """Возвращает фиксированный курс, принимая валюту, написанную по-русски.
                :param currency: Валюта на русском языке.
                """
        return self.currency_to_rub[currency]


class UserInterface:
    """Класс обработки ввода пользовательских данных.
        Attributes
        ----------
        file_name : str
            Путь до CSV-файла.
        profession_name : str
            Название профессии, введённое пользователем.
        """
    file_name: str
    profession_name: str

    def __init__(self, file_name: str = None):
        """Инициализирует объект UserInterface, принимает название CSV-файла.
                :param file_name: Путь до CSV-файла.
                """
        if file_name is not None:
            self.file_name = file_name
        else:
            self.file_name = "../vacancies.csv"
        self.profession_name = 'Аналитик'


class Salary:
    """Класс для предоставления зарплаты.
        Attributes
        ----------
        salary_from : int
            Нижняя граница вилки оклада
        salary_to : int
            Верхняя граница вилки оклада
        salary_currency : str
            Валюта оклада на русском языке
        """
    salary_from: int
    salary_to: int
    salary_currency: str

    def SetField(self, key: str, value: str):
        """Устанавливает поле зарплаты, значение по ключу.
                    :param key: Название поля.
                    :param value: Значение поля. Валюта переводится из международного формата на русский язык.
                        Числовые значения приводятся к int.
                """
        if key == 'salary_currency':
            value = translator.Translate(value)
        if key in ['salary_from', 'salary_to']:
            value = float(value)
        self.__setattr__(key, value)

    def GetAverageInRur(self) -> int:
        """Вычисляет среднюю зарплату из вилки и переводит в рубли при помощи словаря - currency_to_rub.
                Returns:
                    int: Средняя зарплата в рублях
                """
        return int(translator.TranslateCurrencyToRub(self.salary_currency) *
                   (float(self.salary_from) + float(self.salary_to)) // 2)


class Vacancy:
    """Класс вакансии используется для обработки данных о вакансиях из CSV-файлов.
        Attributes
        ----------
        name : str
            Название вакансии
        salary : Salary
            Вилка и валюта оклада
        area_name : str
            Название населённого пункта
        published_at : int
            Время публикации в формате - год.
        """
    name: str
    salary: Salary
    area_name: str
    published_at: int

    def __init__(self, fields: dict):
        """Инициализирует класс вакансии, используя переданные поля.
                :param fields: Словарь с полями вакансии. Доступные ключи - name, salary_from, salary_to, salary_currency,
                area_name, published_at
                """
        for key, value in fields.items():
            if not self.CheckSalary(key, value):
                self.__setattr__(key, self.GetCorrectField(key, value))

    def GetField(self, field: str):
        """Возвращает значение поля вакансии по ключу.
                :param field: Название поля вакансии.
                """
        if field in 'salary':
            return self.salary.GetAverageInRur()
        return self.__getattribute__(field)

    def CheckSalary(self, key: str, value: str) -> bool:
        """Проверяет и устанавливает поле Salary, если его ещё нет
                :param key: Название поля зарплаты, такое как salary_from, salary_to, salary_currency.
                :param value: Значение поля зарплаты, числовое значение или международный формат валюты.
                :returns: Возвращает True, если название поля относится к зарплате.
                """
        is_salary = False
        if key in ['salary_from', 'salary_to', 'salary_currency']:
            if not hasattr(self, 'salary'):
                self.salary = Salary()
            self.salary.SetField(key, value)
            is_salary = True
        return is_salary

    @staticmethod
    def GetCorrectField(key: str, value: str or list) -> int or str:
        """Возвращает отформатированное поле вакансии. Сейчас форматирует только поле published_at.
                :param key: Название поля вакансии.
                :param value: Значение поля вакансии. Дату в формате YY-MM-DDTHH:MM:SS+MS преобразует в год в числовом формате.
                """
        if key == 'published_at':
            big, small = value[:19].split('T')
            year, month, day = big.split('-')
            return int(year)
        else:
            return value


class Report:
    """Класс формирования отчёта по данным из DataSet.
        Attributes
        ----------
        workbook : Workbook
            Класс, содержащий в себе функционал для работы с Excel таблицей.
        data : dict
            Словарь данных, получаемый из DataSet.
        """
    workbook: Workbook
    data: dict

    def __init__(self, data: dict, **kwargs):
        """Инициализирует объект Report. Создаёт пустой Workbook, распаковывает kwargs.
                :param data: Словарь с данными из DataSet.
                """
        self.workbook = Workbook()
        self.data = data
        for key, value in kwargs.items():
            self.__setattr__(key, value)

    # region Excel
    def GenerateExcel(self, file_name: str) -> None:
        """Генерирует и сохраняет Excel-файл.
                :param file_name: название Excel-файла с явно указанным расширением.
                """
        self.FillWithStatistics()
        self.workbook.save(file_name)

    def FillWithStatistics(self) -> None:
        """Заполняет два листа Excel-файла статистикой."""
        self.FillSalariesStatistics()
        self.FillCitiesStatistics()

    def FillSalariesStatistics(self) -> None:
        """Заполняет первую страницу данными о годах, зарплатах и количествах вакансий.
                """
        ws = self.workbook.active
        ws.title = 'Статистика по годам'
        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]

        self.FillColumn('Год', list(salaries_by_years.keys()),
                        [cell[0] for cell in ws['A1':f'A{len(salaries_by_years) + 1}']])

        self.FillColumn('Средняя зарплата', list(salaries_by_years.values()),
                        [cell[0] for cell in ws['B1':f'B{len(salaries_by_years) + 1}']])
        self.FillColumn(f'Средняя зарплата - {ds.profession_name}', list(profession_salaries_by_years.values()),
                        [cell[0] for cell in ws['C1':f'C{len(profession_salaries_by_years) + 1}']])

        self.FillColumn('Количество вакансий', list(vacancies_by_years.values()),
                        [cell[0] for cell in ws['D1':f'D{len(vacancies_by_years) + 1}']])
        self.FillColumn(f'Количество вакансий - {ds.profession_name}', list(profession_vacancies_by_years.values()),
                        [cell[0] for cell in ws['E1':f'E{len(profession_vacancies_by_years) + 1}']])

        self.UpdateWorksheetSettings(ws)

    def FillCitiesStatistics(self) -> None:
        """Создаёт и переключается на второй лист Excel-файла. Заполняет его данными о городах и зарплатах."""
        self.workbook.create_sheet("Статистика по городам")
        ws = self.workbook["Статистика по городам"]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        vacs_ratio_by_cities = self.data["Доля вакансий по городам"]

        self.FillColumn('Город', list(salaries_by_cities.keys()),
                        [cell[0] for cell in ws['A1':f'A{len(salaries_by_cities) + 1}']])
        self.FillColumn('Уровень зарплат', list(salaries_by_cities.values()),
                        [cell[0] for cell in ws['B1': f'B{len(salaries_by_cities) + 1}']])

        self.FillColumn('Город', list(vacs_ratio_by_cities.keys()),
                        [cell[0] for cell in ws['D1':f'D{len(vacs_ratio_by_cities) + 1}']])
        self.FillColumn('Доля вакансий', list(vacs_ratio_by_cities.values()),
                        [cell[0] for cell in ws['E1': f'E{len(vacs_ratio_by_cities) + 1}']])

        self.SetColumnPercent([cell[0] for cell in ws['E2': f'E{len(vacs_ratio_by_cities) + 1}']])
        self.UpdateWorksheetSettings(ws)

    @staticmethod
    def FillColumn(header: str, data: list, column_cells: list) -> None:
        """Заполняет столбец данными.
                :param header: Заголовок столбца, записывается в первой ячейке из списка клеток.
                :param data: Данные для записи в клетки.
                :param column_cells: Список клеток, в которые будут записаны данные.
                """
        column_cells[0].value = header
        for cell, value in zip(column_cells[1:], data):
            cell.value = value

    @staticmethod
    def SetColumnPercent(column: list) -> None:
        """Устанавливает процентный формат для всех ячеек в этом столбце.
                """
        for cell in column:
            cell.number_format = FORMAT_PERCENTAGE_00

    def UpdateWorksheetSettings(self, ws) -> None:
        """Устанавливает границы и ширины для страницы Excel-файла.
                :param ws: страница Excel-файла.
                """
        self.SetBorders(ws)
        self.SetColumnWidth(ws)

    @staticmethod
    def SetBorders(ws) -> None:
        """Устанавливает границы для всех непустых клеток и жирный шрифт для заголовков столбцов.
                :param ws: страница Excel-файла.
                """
        isFirstRow = True
        for row in ws.rows:
            for cell in row:
                if not cell.value:
                    continue
                cell.border = Border(top=Side(border_style="thin", color="000000"),
                                     left=Side(border_style="thin", color="000000"),
                                     right=Side(border_style="thin", color="000000"),
                                     bottom=Side(border_style="thin", color="000000"))
                if isFirstRow:
                    cell.font = Font(bold=True)
            isFirstRow = False

    @staticmethod
    def SetColumnWidth(ws) -> None:
        """Устанавливает ширину столбца, ориентируясь на максимально большую ячейку в нём.
                :param ws: страница Excel-файла.
                """
        a = {0: "A", 1: "B", 2: "C", 3: "D", 4: "E", 6: "F", 7: "G"}
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value)) + 1))

        for col, value in dims.items():
            ws.column_dimensions[a[col - 1]].width = value

    # endregion
    # region Plot

    def GenerateImage(self, file_name: str, show_result: bool = False) -> None:
        """Генерирует и сохраняет изображение с графиками, на основе данных data.
                :param file_name: Название для сохранения изображения.
                :param show_result: Показывать ли изображение после генерации. По-умолчанию False.
                """
        self.DrawGraphs()
        plt.tight_layout()
        plt.savefig(file_name, dpi=300)
        if show_result:
            plt.show()

    def DrawGraphs(self) -> None:
        """Рисует 4 графика на сетке 2x2. Каждый график строится на основании данных каждого ключа из data.
                """
        figure, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        self.DrawBarGraph(ax1, "Уровень зарплат по годам")
        self.DrawBarGraph(ax2, "Количество вакансий по годам")
        self.DrawInvertBarGraph(ax3, "Уровень зарплат по городам")
        self.DrawPieGraph(ax4, "Доля вакансий по городам")

    def DrawBarGraph(self, subplot, name: str) -> None:
        """Рисует столбчатую диаграмму.
                :param subplot: Подобласть для отрисовки графика.
                :param name: Название графика. Должен соответствовать ключу из data.
                """
        bar_width = 0.4
        first_label = 'средняя з/п'
        second_label = f'з/п {ui.profession_name}'
        if name == "Количество вакансий по годам":
            first_label = "Количество вакансий"
            second_label = f"Количество вакансий\n{ui.profession_name}"

        average_by_years: dict = self.data[name][0]
        profession_average_by_years: dict = self.data[name][1]

        X_axis = np.arange(len(average_by_years.keys()))

        subplot.bar(X_axis - bar_width / 2, average_by_years.values(), width=bar_width, label=first_label)
        subplot.bar(X_axis + bar_width / 2, profession_average_by_years.values(),
                    width=bar_width, label=second_label)
        subplot.set_xticks(X_axis, average_by_years.keys())
        subplot.set_xticklabels(average_by_years.keys(), rotation='vertical', va='top', ha='center')

        subplot.set_title(name)
        subplot.grid(True, axis='y')
        subplot.tick_params(axis='both', labelsize=8)
        subplot.legend(fontsize=8)

    def DrawInvertBarGraph(self, subplot, name: str) -> None:
        """Рисует повёрнутую на левый бок столбчатую диаграмму.
                :param subplot: Подобласть для отрисовки графика.
                :param name: Название графика. Должен соответствовать ключу из data.
                """
        subplot.invert_yaxis()
        courses = list(self.data[name].keys())
        courses = [label.replace(' ', '\n').replace('-', '-\n') for label in courses]
        values = list(self.data[name].values())
        subplot.barh(courses, values)
        subplot.set_yticklabels(courses, fontsize=6, va='center', ha='right')

        subplot.set_title(name)
        subplot.grid(True, axis='x')
        subplot.tick_params(axis='both', labelsize=8)

    def DrawPieGraph(self, subplot, name: str) -> None:
        """Рисует круговую диаграмму.
                :param subplot: Подобласть для отрисовки графика.
                :param name: Название графика. Должен соответствовать ключу из data.
                """
        data = self.data[name]
        other = 1 - sum((list(data.values())))
        new_dic = {'Другие': other}
        new_dic.update(data)

        labels = list(new_dic.keys())
        sizes = list(new_dic.values())

        subplot.set_title(name)
        subplot.pie(sizes, labels=labels, textprops={'fontsize': 6})
        subplot.axis('scaled')

    # endregion
    # region PDF

    def GeneratePdf(self, name: str):
        """Генерирует PDF-файл на основании данных из DataSet - data и разметки PDF - pdf_template.html.
                :param name: Название сохраняемого PDF-файла с явно указанным расширением.
                """
        image_file = "graph.png"
        header_year = ["Год", "Средняя зарплата", f"Средняя зарплата - {ds.profession_name}", "Количество вакансий",
                       f"Количество вакансий - {ds.profession_name}"]
        header_city = ["Город", "Уровень зарплат", '', "Город", "Доля вакансий"]

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        salaries_by_years = self.data["Уровень зарплат по годам"][0]
        vacancies_by_years = self.data["Количество вакансий по годам"][0]
        profession_salaries_by_years = self.data["Уровень зарплат по годам"][1]
        profession_vacancies_by_years = self.data["Количество вакансий по годам"][1]
        salaries_by_cities = self.data["Уровень зарплат по городам"]
        ratio_vacancy_by_cities = {city: str(f'{ratio * 100:,.2f}%').replace('.', ',')
                                   for city, ratio in self.data["Доля вакансий по городам"].items()}

        salary_data = {year: [salary, count, salary_vac, count_vac]
                       for year, salary, count, salary_vac, count_vac in zip(salaries_by_years.keys(),
                                                                             salaries_by_years.values(),
                                                                             vacancies_by_years.values(),
                                                                             profession_salaries_by_years.values(),
                                                                             profession_vacancies_by_years.values())}
        city_data = {index: [salary_city, salary, ratio_city, ratio]
                     for index, (salary_city, salary, ratio_city, ratio) in
                     enumerate(zip(salaries_by_cities.keys(),
                                   salaries_by_cities.values(),
                                   ratio_vacancy_by_cities.keys(),
                                   ratio_vacancy_by_cities.values()))}

        pdf_template = template.render(
            {'image_file': image_file,
             'image_style': 'style="max-width:1024px; max-height:680px"',
             'salary_data': salary_data,
             'city_data': city_data,
             'header_year': header_year,
             'header_city': header_city,
             'profession_name': f"{ui.profession_name}",
             'h1_style': 'style="text-align:center; font-size:32px"',
             'h2_style': 'style="text-align:center"',
             'cell_style_none': "style=''",
             'cell_style': 'style="border:1px solid black; border-collapse: collapse; font-size: 16px; height: 19pt;'
                           'padding: 5px; text-align:center"'})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, name, configuration=config, options={'enable-local-file-access': None})
    # endregion


class DataSet:
    """Класс хранилища данных о вакансиях.
        Attributes
        ----------
        profession_name : str
            Название профессии, введённой пользователем.
        profession_count : int
            Количество профессий, содержащих в своём названии profession_name.
        vacancies : List[Vacancy]
            Список вакансий.
        salary_by_years : {int, list}
            Год: средняя зарплата среди всех вакансий за этот период.
        vacancies_by_years : {int, int}
            Год: общее количество вакансий за этот период.
        profession_salary_by_years : {int, list}
            Год: средняя зарплата среди вакансий, содержащих в своём названии profession_name, за этот период.
        profession_vacancies_by_years : {int, int}
            Год: количество вакансий, содержащих в своём названии profession_name, за этот период.
        salaries_by_cities : {str, list}
            Название города: [сумма всех зарплат вакансий в этом городе, количество вакансий в этом городе].
        ratio_vacancy_by_cities : {str, float}
            Название города: доля количества вакансий в этом городе к общему количеству вакансий.
        city_vacancies_count : {str, int}
            Название города: количество вакансий в этом городе.
        """
    profession_name: str
    profession_count: int
    vacancies: List[Vacancy]
    salary_by_years: {int, list}
    vacancies_by_years: {int, int}
    profession_salary_by_years: {int, list}
    profession_vacancies_by_years: {int, int}
    salaries_by_cities: {str, list}
    ratio_vacancy_by_cities: {str, float}
    city_vacancies_count: {str, int}

    def __init__(self, vacs: list, prof_name: str):
        self.profession_name = prof_name
        self.profession_count = 0
        self.vacancies = vacs
        self.salary_by_years = {}
        self.vacancies_by_years = {}
        self.profession_salary_by_years = {}
        self.profession_vacancies_by_years = {}
        self.salaries_by_cities = {}
        self.ratio_vacancy_by_cities = {}
        self.city_vacancies_count = {}

        self._GetData()

    def _GetData(self) -> None:
        """Обрабатывает данные вакансий из инициализированного списка"""
        for vac in self.vacancies:
            self.ProcessVacanciesCount('city_vacancies_count', 'area_name', vac)
        for vac in self.vacancies:
            self.ProcessSalary('salary_by_years', 'published_at', vac)
            self.ProcessVacanciesCount('vacancies_by_years', 'published_at', vac)
            if self.profession_name in vac.name:
                self.profession_count += 1
                self.ProcessSalary('profession_salary_by_years', 'published_at', vac)
                self.ProcessVacanciesCount('profession_vacancies_by_years', 'published_at', vac)
            self.ProcessSalary('salaries_by_cities', 'area_name', vac)
            self.ProcessVacanciesCount('ratio_vacancy_by_cities', 'area_name', vac)

        self.SetCorrectCitiesData()

    def ProcessSalary(self, dict_name: str, field: str, vac: Vacancy) -> None:
        """Обрабатывает поля, связанные с зарплатой вакансии, заполняя словарь, имя которого было передано.
                :param dict_name: Название словаря этого объекта.
                :param field: Поле вакансии, которое будет обрабатываться.
                :param vac: Экземпляр вакансии.
                :return:
                """
        d = self.__getattribute__(dict_name)
        f = vac.GetField(field)
        if f not in d.keys():
            d[f] = [vac.salary.GetAverageInRur(), 1]
        else:
            d[f][0] += vac.salary.GetAverageInRur()
            d[f][1] += 1

    def SetCorrectCitiesData(self):
        """Обрабатывает словари, связанные с данными по городам. Сортирует словари по значениям - средней зарплате
                и доле вакансии в городе. Наибольшие значения идут первыми.
                """
        for key, value in self.ratio_vacancy_by_cities.items():
            self.ratio_vacancy_by_cities[key] = round(value / len(self.vacancies), 4)

        d1 = dict(sorted(self.salaries_by_cities.items(), key=lambda i: i[1][1] / i[1][0]))
        self.salaries_by_cities = self.GetFirstTenCorrect(d1)

        d2 = dict(sorted(self.ratio_vacancy_by_cities.items(), key=lambda i: i[1], reverse=True))
        self.ratio_vacancy_by_cities = self.GetFirstTenCorrect(d2)

    def ProcessVacanciesCount(self, dict_name: str, field: str, vac: Vacancy) -> None:
        """Обрабатывает поля, связанные с количеством вакансий, заполняя словарь, имя которого было передано.
                :param dict_name: Название словаря этого объекта.
                :param field: Поле вакансии, которое будет обрабатываться.
                :param vac: Экземпляр вакансии.
                :return:
                """
        d = self.__getattribute__(dict_name)
        f = vac.GetField(field)
        if f not in d.keys():
            d[f] = 1
        else:
            d[f] += 1

    def GetFirstTenCorrect(self, d: dict) -> dict:
        """Оставляет в словаре только первые 10 значений, удовлетворяющих условию - больше 1% вакансий в городе от
                общего числа вакансий."""
        count = 0
        res = {}
        for key, value in d.items():
            if count == 10:
                break
            if self.city_vacancies_count[key] >= len(self.vacancies) // 100:
                res[key] = value
                count += 1
        return res

    def GetData(self) -> dict:
        """Обрабатывает полученные данные.
                :returns: "Уровень зарплат по годам": {год: средняя зарплата за этот период},
                          "Количество вакансий по годам": {год: общее количество вакансий за этот период},
                          "Уровень зарплат по городам": {город: средняя зарплата},
                          "Доля вакансий по городам": {доля вакансий от общего количества вакансий}.
                          Для статистики по городам возвращается только 10 городов с наибольшими значениями."""
        salaries_by_years, vacancies_by_years = [], []
        salaries_by_cities, ratio_vacancies_by_cities = {}, {}
        to_print: {str, dict} \
            = {"Уровень зарплат по годам": self.salary_by_years,
               "Количество вакансий по годам": self.vacancies_by_years,
               "Уровень зарплат по годам для выбранной профессии": self.profession_salary_by_years,
               "Количество вакансий по годам для выбранной профессии": self.profession_vacancies_by_years,
               "Уровень зарплат по городам": self.salaries_by_cities,
               "Доля вакансий по городам": self.ratio_vacancy_by_cities}
        for key, value in to_print.items():
            if len(value) == 0:
                value = {k: 0 for k in self.salary_by_years.keys()}
            for k, v in value.items():
                if type(v) is list:
                    value[k] = v[0] // v[1]
            if 'Уровень зарплат по годам' in key:
                salaries_by_years.append(value)
            elif 'Количество вакансий по годам' in key:
                vacancies_by_years.append(value)
            elif 'Уровень зарплат по городам' in key:
                salaries_by_cities = value
            else:
                ratio_vacancies_by_cities = value

        return {"Уровень зарплат по годам": salaries_by_years,
                "Количество вакансий по годам": vacancies_by_years,
                "Уровень зарплат по городам": salaries_by_cities,
                "Доля вакансий по городам": ratio_vacancies_by_cities}


def GetParsedRowVacancy(row_vacs: list) -> dict:
    """Очищает строки от HTML-тегов и разбивает её на данные для вакансии.
        :param header: список заголовков из CSV-файла.
        :param row_vacs: список строк, прочитанных из CSV-файла.
        """
    return dict(zip(title, map(GetParsedString, row_vacs)))


def GetParsedString(line: str) -> str:
    """Убирает HTML-теги из строки.
        :param line: Строка для обработки.
        :returns: Возвращает строку без HTML-тегов.
        """
    line = sub('<.*?>', '', line)
    res = [' '.join(word.split()) for word in line.replace("\r\n", "\n").split('\n')]
    return res[0] if len(res) == 1 else res


if __name__ == '__main__':
    translator = Translator()
    ui = UserInterface("vacancies.csv")
    csv = CSV(ui.file_name)
    title, row_vacancies = csv.title, csv.rows
    vacancies = [Vacancy(GetParsedRowVacancy(row_vac)) for row_vac in row_vacancies]
    ds = DataSet(vacancies, ui.profession_name)
    statistics = ds.GetData()
    report = Report(statistics)
    report.GeneratePdf('report.pdf')
    report.GenerateExcel('repo rt.xlsx')
    report.GenerateImage('graph.png')
